#!/usr/bin/env python3
from openpyxl import load_workbook
import sys, matplotlib.pyplot as plt, matplotlib.patches as mpatches, numpy as np, matplotlib.pyplot as mpl

def printPoints(sheet, firstRow, firstCol, lastCol, maxClubs, gamesPlayed):
    pointsNew = [None] * gamesPlayed
    rangePoints = [None] * gamesPlayed
    for row in range(firstRow, maxClubs + firstRow):
        for value in sheet.iter_rows(min_row = row, max_row = row, min_col = firstCol,
                                    max_col = lastCol , values_only = True):
            if gamesPlayed == 1:
                points = value[0]
            else:
                points = value[0:gamesPlayed]

            print(points)
            checkGames(gamesPlayed)
            createLists(points, pointsNew, rangePoints, gamesPlayed)
            createGraphic(pointsNew, rangePoints, gamesPlayed)

def checkGames(gamesPlayed):
    if gamesPlayed == 1:
        print("Games played is equal to one, can't graphic it")
        sys.exit(1)

def createLists(points, pointsNew, rangePoints, gamesPlayed):
    temp = 0
    for i in range(0, gamesPlayed):
        pointsNew[i] = points[i] - 1.5 + temp
        temp = pointsNew[i]
    for j in range(1, gamesPlayed + 1):
        rangePoints[j - 1] = j

def createGraphic(pointsNew, rangePoints, gamesPlayed):

    mpl.rcParams['toolbar'] = 'None'
    fig, ax = plt.subplots(figsize = (10, 2))
    fig.tight_layout()
    fig.patch.set_facecolor('xkcd:gray')
    ax.scatter(rangePoints, pointsNew, color = 'b')
    ax.plot(rangePoints, pointsNew, color = 'k')
    ax.set_xlim(0, gamesPlayed)
    ax.set_facecolor('xkcd:gray')
    ax.set_yticklabels([])
    ax.set_ylabel('Puntuación', fontsize = 10)
    ax.set_xlabel('Jornadas', fontsize = 10)
    ax.set_title('Club\n')
    ax.grid(linestyle = '--', linewidth = 0.5)
    plt.show()

def main():
    workbook = load_workbook(filename = "Quiniela.xlsx")
    workbook.sheetnames
    sheet = workbook.active

    maxClubsSantander = 20
    maxClubsSmartbank = 22
    firstRowSantander = 3
    firstRowSmartbank = 25
    firstCol = 6
    lastCol = 43

    gamesPlayed = 38

    print("Liga Santander")
    printPoints(sheet, firstRowSantander, firstCol, lastCol, maxClubsSantander, gamesPlayed)
    print("\nLiga Smartbank")
    printPoints(sheet, firstRowSmartbank, firstCol, lastCol, maxClubsSmartbank, gamesPlayed)

if __name__ == "__main__":
    main()
