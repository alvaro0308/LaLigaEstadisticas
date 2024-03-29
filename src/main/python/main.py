"""LaLigaEstadisticas App"""

# export PYTHONIOENCODING=utf-8

import os
import sys
import yaml
from AppUI import AppUI
from PyQt5 import QtCore
from PyQt5.QtCore import Qt
from ConfigApp import ConfigApp
from openpyxl import load_workbook
from PyQt5.QtGui import QPalette, QColor
from PyQt5.QtWidgets import QApplication, QMessageBox
from fbs_runtime.application_context.PyQt5 import ApplicationContext


__version__ = "alpha"
__author__ = "Alvaro"


class App:
    """LaLigaEstadisticas App"""

    def __init__(self, path):
        """Read clubs and launch Application"""
        with open(path) as f:
            self.params = yaml.load(f, Loader=yaml.FullLoader)
        path = path.replace("config/config.yaml", "")
        self.params['path'] = path

        workbook = load_workbook(
            filename=self.params['path'] + self.params['databasePath']
            + self.params['nameDownloadedDatabase'] + self.params['extensionDatabase'])
        self.sheet = workbook.active

        maxClubsSantander = self.params['maxClubsSantander']
        maxClubsSmartbank = self.params['maxClubsSmartbank']
        firstRowSantander = self.params['firstRowSantander']
        firstRowSmartbank = self.params['firstRowSmartbank']

        listClubsSantander = []
        listClubsSmartbank = []
        dictMistersSantander = {}
        dictMistersSmartbank = {}
        dictCommentsSantander = {}
        dictCommentsSmartbank = {}
        dictStandingSantander = {}
        dictStandingSmartbank = {}

        self.readClubs(listClubsSantander, maxClubsSantander,
                       firstRowSantander)
        self.readClubs(listClubsSmartbank, maxClubsSmartbank,
                       firstRowSmartbank)

        self.readMisters(dictMistersSantander, maxClubsSantander,
                         firstRowSantander)
        self.readMisters(dictMistersSmartbank, maxClubsSmartbank,
                         firstRowSmartbank)

        for clubSantander in listClubsSantander:
            dictCommentsSantander[clubSantander] = self.readCommentsClubs(
                clubSantander)
        for clubSmartbank in listClubsSmartbank:
            dictCommentsSmartbank[clubSmartbank] = self.readCommentsClubs(
                clubSmartbank)

        self.readStandings(
            listClubsSantander, maxClubsSantander, firstRowSantander, dictStandingSantander, self.params['lastColPointsSantander'] - 1)
        self.readStandings(
            listClubsSmartbank, maxClubsSmartbank, firstRowSmartbank, dictStandingSmartbank, self.params['lastColPointsSmartbank'] + 1)

        dictStandingSantander = dict(
            sorted(dictStandingSantander.items(), key=lambda item: item[1], reverse=True))
        dictStandingSmartbank = dict(
            sorted(dictStandingSmartbank.items(), key=lambda item: item[1], reverse=True))

        self.app = QApplication(sys.argv)
        # msg = QMessageBox()
        # msg.setIcon(QMessageBox.Critical)
        # msg.setText("Error")
        # msg.setInformativeText(
        #    'IndexError: list index out of range club gamesHome')
        # msg.setWindowTitle("Error")
        # msg.exec_()

        os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
        os.environ["QT_ENABLE_HIGHDPI_SCALING"] = "1"
        self.app.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)
        self.app.setAttribute(
            QtCore.Qt.AA_UseHighDpiPixmaps, True)
        self.darkMode()
        view = AppUI(self.sheet, listClubsSantander, listClubsSmartbank,
                     dictMistersSantander, dictMistersSmartbank,
                     dictCommentsSantander, dictCommentsSmartbank,
                     maxClubsSantander, maxClubsSmartbank,
                     firstRowSantander, firstRowSmartbank,
                     dictStandingSantander, dictStandingSmartbank, self.params)
        view.show()
        sys.exit(self.app.exec_())

    def readClubs(self, listClubs, maxClubs, firstRow):
        """Read clubs from database"""
        for cell in self.sheet.iter_rows(min_row=firstRow,
                                         max_row=maxClubs + firstRow - 1,
                                         min_col=5, max_col=5,
                                         values_only=True):
            listClubs.append(cell[0])

    def readStandings(self, listClubs, maxClubs, firstRow, dictStandings, col):
        """Read standings from database"""
        index = 0
        for cell in self.sheet.iter_rows(min_row=firstRow,
                                         max_row=maxClubs + firstRow - 1,
                                         min_col=col, max_col=col,
                                         values_only=True):
            try:
                points = cell[0].replace(',', '.')
                dictStandings[listClubs[index]] = float(points)
            except Exception as exception:
                points = []
                print("Exception: {} points: {} ".format(
                    exception, points))
            index += 1

    def readMisters(self, dictMisters, maxClubs, firstRow):
        """Read misters from database"""
        listNamesMister = []
        for cell in self.sheet.iter_rows(min_row=firstRow,
                                         max_row=maxClubs + firstRow - 1,
                                         min_col=5, max_col=5,
                                         values_only=True):
            listNamesMister.append(cell[0])

        currentMister = 0
        for cell in self.sheet.iter_rows(min_row=firstRow,
                                         max_row=maxClubs + firstRow - 1,
                                         min_col=6, max_col=6,
                                         values_only=True):
            dictMisters[listNamesMister[currentMister]] = cell[0]
            currentMister += 1

    def readCommentsClubs(self, club):
        """Read comments clubs from database"""
        counterBlank = 0
        currentRow = 3
        listComments = []
        for cell in self.sheet.iter_rows(min_row=3,
                                         min_col=2, max_col=2,
                                         values_only=True):

            if cell[0] is None:
                counterBlank += 1
            else:
                counterBlank = 0
            # if counterBlank > 2:
            #     break
            clubs = str(self.sheet.cell(currentRow, 1).value).split(" - ")
            for i in clubs:
                if i == club:
                    listComments.append(str(self.sheet.cell(currentRow, 1).value) + ": "
                                        + str(cell[0]) + "_ "
                                        + str(self.sheet.cell(currentRow, 3).value))
            currentRow += 1

        return listComments

    def darkMode(self):
        """Apply dark mode to Application"""
        self.app.setStyle('Fusion')
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(53, 53, 53))
        palette.setColor(QPalette.WindowText, Qt.white)
        palette.setColor(QPalette.Base, QColor(15, 15, 15))
        palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        palette.setColor(QPalette.ToolTipBase, Qt.white)
        palette.setColor(QPalette.ToolTipText, Qt.white)
        palette.setColor(QPalette.Text, Qt.white)
        palette.setColor(QPalette.Button, QColor(53, 53, 53))
        palette.setColor(QPalette.ButtonText, Qt.white)
        palette.setColor(QPalette.BrightText, Qt.red)
        palette.setColor(QPalette.Highlight, QColor(0, 87, 184).lighter())
        palette.setColor(QPalette.HighlightedText, Qt.black)
        self.app.setPalette(palette)


def main():
    """Is main function"""
    configApp = ConfigApp()
    App(configApp.path)


if __name__ == '__main__':
    main()
